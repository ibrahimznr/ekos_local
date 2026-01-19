from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List
from datetime import datetime, timezone
import io
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models.makine import Makine, MakineCreate, MakineUpdate
from routers.auth import get_current_user
from database import db

router = APIRouter(prefix="/makineler", tags=["Makineler"])

@router.get("", response_model=List[Makine])
async def get_makineler(current_user: dict = Depends(get_current_user)):
    """Tüm makineleri listele"""
    makineler = await db.makineler.find({}, {"_id": 0}).to_list(1000)
    for makine in makineler:
        if isinstance(makine.get('created_at'), str):
            makine['created_at'] = datetime.fromisoformat(makine['created_at'])
        if isinstance(makine.get('updated_at'), str):
            makine['updated_at'] = datetime.fromisoformat(makine['updated_at'])
    return makineler

@router.get("/{makine_id}", response_model=Makine)
async def get_makine(makine_id: str, current_user: dict = Depends(get_current_user)):
    """Tek bir makineyi ID ile getir"""
    makine = await db.makineler.find_one({"id": makine_id}, {"_id": 0})
    if not makine:
        raise HTTPException(status_code=404, detail="Makine bulunamadı")
    if isinstance(makine.get('created_at'), str):
        makine['created_at'] = datetime.fromisoformat(makine['created_at'])
    if isinstance(makine.get('updated_at'), str):
        makine['updated_at'] = datetime.fromisoformat(makine['updated_at'])
    return makine

@router.post("", response_model=Makine)
async def create_makine(makine_create: MakineCreate, current_user: dict = Depends(get_current_user)):
    """Yeni makine ekle"""
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    # Aynı plaka/seri numarası var mı kontrol et
    existing = await db.makineler.find_one({"plaka_seri_no": makine_create.plaka_seri_no})
    if existing:
        raise HTTPException(status_code=400, detail="Bu plaka/seri numarası zaten kayıtlı")
    
    makine = Makine(**makine_create.model_dump())
    doc = makine.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('updated_at'):
        doc['updated_at'] = doc['updated_at'].isoformat()
    await db.makineler.insert_one(doc)
    return makine

@router.put("/{makine_id}", response_model=Makine)
async def update_makine(makine_id: str, makine_update: MakineUpdate, current_user: dict = Depends(get_current_user)):
    """Makine güncelle"""
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.makineler.find_one({"id": makine_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Makine bulunamadı")
    
    # Plaka değişiyorsa, başka makine kullanıyor mu kontrol et
    if makine_update.plaka_seri_no and makine_update.plaka_seri_no != existing.get('plaka_seri_no'):
        duplicate = await db.makineler.find_one({"plaka_seri_no": makine_update.plaka_seri_no})
        if duplicate:
            raise HTTPException(status_code=400, detail="Bu plaka/seri numarası zaten kayıtlı")
    
    update_data = makine_update.model_dump(exclude_unset=True)
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.makineler.update_one({"id": makine_id}, {"$set": update_data})
    
    updated_makine = await db.makineler.find_one({"id": makine_id}, {"_id": 0})
    if isinstance(updated_makine.get('created_at'), str):
        updated_makine['created_at'] = datetime.fromisoformat(updated_makine['created_at'])
    if isinstance(updated_makine.get('updated_at'), str):
        updated_makine['updated_at'] = datetime.fromisoformat(updated_makine['updated_at'])
    return updated_makine

@router.delete("/{makine_id}")
async def delete_makine(makine_id: str, current_user: dict = Depends(get_current_user)):
    """Makine sil"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.makineler.delete_one({"id": makine_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Makine bulunamadı")
    return {"message": "Makine silindi"}

@router.post("/bulk-delete")
async def bulk_delete_makineler(makine_ids: List[str], current_user: dict = Depends(get_current_user)):
    """Toplu makine sil"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.makineler.delete_many({"id": {"$in": makine_ids}})
    return {"message": f"{result.deleted_count} makine silindi", "deleted_count": result.deleted_count}

# Excel Endpoints
@router.get("/excel/export")
async def export_makineler_excel(current_user: dict = Depends(get_current_user)):
    """Tüm makineleri Excel'e aktar"""
    makineler = await db.makineler.find({}, {"_id": 0}).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Makineler"
    
    headers = [
        "Makine Türü", "Firma", "Plaka/Seri No", "Şasi/Motor No", "İmalat Yılı",
        "Servis Bakım Tarihi", "Sigorta Tarihi", "Periyodik Kontrol Tarihi",
        "Ruhsat Muayene Tarihi", "Operatör Adı", "Operatör Belge Tarihi",
        "Belge Kurumu", "Telefon", "Proje", "Durum", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, makine in enumerate(makineler, 2):
        ws.cell(row=row_idx, column=1, value=makine.get("makine_turu", ""))
        ws.cell(row=row_idx, column=2, value=makine.get("firma", ""))
        ws.cell(row=row_idx, column=3, value=makine.get("plaka_seri_no", ""))
        ws.cell(row=row_idx, column=4, value=makine.get("sasi_motor_no", ""))
        ws.cell(row=row_idx, column=5, value=makine.get("imalat_yili", ""))
        ws.cell(row=row_idx, column=6, value=makine.get("servis_bakim_tarihi", ""))
        ws.cell(row=row_idx, column=7, value=makine.get("sigorta_tarihi", ""))
        ws.cell(row=row_idx, column=8, value=makine.get("periyodik_kontrol_tarihi", ""))
        ws.cell(row=row_idx, column=9, value=makine.get("ruhsat_muayene_tarihi", ""))
        ws.cell(row=row_idx, column=10, value=makine.get("operator_adi", ""))
        ws.cell(row=row_idx, column=11, value=makine.get("operator_belge_tarihi", ""))
        ws.cell(row=row_idx, column=12, value=makine.get("belge_kurumu", ""))
        ws.cell(row=row_idx, column=13, value=makine.get("telefon", ""))
        ws.cell(row=row_idx, column=14, value=makine.get("proje_adi", ""))
        ws.cell(row=row_idx, column=15, value=makine.get("durum", ""))
        ws.cell(row=row_idx, column=16, value=makine.get("aciklama", ""))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=makineler_{len(makineler)}_adet.xlsx"}
    )

@router.get("/excel/template")
async def download_makine_template():
    """Makine Excel şablonunu indir"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Makine Şablonu"
    
    headers = [
        "Makine Türü", "Firma", "Plaka/Seri No", "Şasi/Motor No", "İmalat Yılı",
        "Servis Bakım Tarihi", "Sigorta Tarihi", "Periyodik Kontrol Tarihi",
        "Ruhsat Muayene Tarihi", "Operatör Adı", "Operatör Belge Tarihi",
        "Belge Kurumu", "Telefon", "Durum", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    examples = [
        ["Forklift", "ABC Firma", "01 AHC 823", "CH123456", "2020", "2025-06-15", "2025-12-31",
         "2025-09-30", "2025-11-15", "Ahmet Yılmaz", "2024-01-10", "Ekipler", "5551234567", "Aktif", ""],
        ["Vinç", "XYZ Ltd", "34 DEF 456", "", "2019", "", "", "", "", "", "", "", "", "Aktif", ""],
        ["Traktör", "Test Firma", "06 GHI 789", "TR987654", "2021", "", "", "", "", "", "", "", "", "Pasif", "Bakımda"]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=makine_sablonu.xlsx"}
    )

@router.post("/excel/import")
async def import_makineler_excel(
    file: UploadFile = File(...),
    proje_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Excel'den toplu makine içe aktar"""
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Excel içe aktarma yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    proje_adi = proje.get("proje_adi", "")
    
    content = await file.read()
    excel_file = io.BytesIO(content)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            
            try:
                def get_cell(index):
                    return row[index] if index < len(row) and row[index] is not None else None
                
                makine_turu = str(get_cell(0)) if get_cell(0) else ""
                firma = str(get_cell(1)) if get_cell(1) else ""
                plaka_seri_no = str(get_cell(2)) if get_cell(2) else ""
                sasi_motor_no = str(get_cell(3)) if get_cell(3) else None
                imalat_yili = str(get_cell(4)) if get_cell(4) else None
                servis_bakim_tarihi = str(get_cell(5)) if get_cell(5) else None
                sigorta_tarihi = str(get_cell(6)) if get_cell(6) else None
                periyodik_kontrol_tarihi = str(get_cell(7)) if get_cell(7) else None
                ruhsat_muayene_tarihi = str(get_cell(8)) if get_cell(8) else None
                operator_adi = str(get_cell(9)) if get_cell(9) else None
                operator_belge_tarihi = str(get_cell(10)) if get_cell(10) else None
                belge_kurumu = str(get_cell(11)) if get_cell(11) else None
                telefon = str(get_cell(12)) if get_cell(12) else None
                durum = str(get_cell(13)) if get_cell(13) else "Aktif"
                aciklama = str(get_cell(14)) if get_cell(14) else None
                
                if not makine_turu or not firma:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik (Makine Türü, Firma)")
                    continue
                
                # Plaka kontrolü (eğer varsa)
                if plaka_seri_no:
                    existing = await db.makineler.find_one({"plaka_seri_no": plaka_seri_no})
                    if existing:
                        errors.append(f"Satır {row_idx}: Bu plaka/seri numarası zaten kayıtlı - '{plaka_seri_no}'")
                        continue
                
                makine_data = {
                    "id": str(uuid.uuid4()),
                    "proje_id": proje_id,
                    "proje_adi": proje_adi,
                    "makine_turu": makine_turu,
                    "firma": firma,
                    "plaka_seri_no": plaka_seri_no,
                    "sasi_motor_no": sasi_motor_no,
                    "imalat_yili": imalat_yili,
                    "servis_bakim_tarihi": servis_bakim_tarihi,
                    "sigorta_tarihi": sigorta_tarihi,
                    "periyodik_kontrol_tarihi": periyodik_kontrol_tarihi,
                    "ruhsat_muayene_tarihi": ruhsat_muayene_tarihi,
                    "operator_adi": operator_adi,
                    "operator_belge_tarihi": operator_belge_tarihi,
                    "belge_kurumu": belge_kurumu,
                    "telefon": telefon,
                    "durum": durum,
                    "aciklama": aciklama,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.makineler.insert_one(makine_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
        
        return {
            "message": f"{imported_count} makine başarıyla içe aktarıldı",
            "imported_count": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel işleme hatası: {str(e)}")
