from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from pathlib import Path
from pydantic import BaseModel
import io
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import Rapor
from routers.auth import get_current_user
from database import db
from utils import generate_rapor_no
from constants import SEHIRLER

router = APIRouter(prefix="/excel", tags=["Excel"])

# Request model for selective export
class ExcelExportRequest(BaseModel):
    rapor_ids: List[str]

@router.post("/export")
async def export_excel_selected(request: ExcelExportRequest, current_user: dict = Depends(get_current_user)):
    """Seçili raporları Excel'e aktar"""
    if not request.rapor_ids:
        raise HTTPException(status_code=400, detail="En az bir rapor seçilmelidir")
    
    raporlar = await db.raporlar.find({"id": {"$in": request.rapor_ids}}, {"_id": 0}).to_list(10000)
    
    if not raporlar:
        raise HTTPException(status_code=404, detail="Seçilen raporlar bulunamadı")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Raporlar"
    
    headers = [
        "Rapor No", "Ekipman Adı", "Kategori", "Firma", "Lokasyon",
        "Marka/Model", "Seri No", "Alt Kategori", "Periyot",
        "Geçerlilik Tarihi", "Uygunluk", "Proje", "Şehir", "Açıklama", "Oluşturma Tarihi"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, rapor in enumerate(raporlar, 2):
        ws.cell(row=row_idx, column=1, value=rapor.get("rapor_no", ""))
        ws.cell(row=row_idx, column=2, value=rapor.get("ekipman_adi", ""))
        ws.cell(row=row_idx, column=3, value=rapor.get("kategori", ""))
        ws.cell(row=row_idx, column=4, value=rapor.get("firma", ""))
        ws.cell(row=row_idx, column=5, value=rapor.get("lokasyon", ""))
        ws.cell(row=row_idx, column=6, value=rapor.get("marka_model", ""))
        ws.cell(row=row_idx, column=7, value=rapor.get("seri_no", ""))
        ws.cell(row=row_idx, column=8, value=rapor.get("alt_kategori", ""))
        ws.cell(row=row_idx, column=9, value=rapor.get("periyot", ""))
        ws.cell(row=row_idx, column=10, value=rapor.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=11, value=rapor.get("uygunluk", ""))
        ws.cell(row=row_idx, column=12, value=rapor.get("proje_adi", ""))
        ws.cell(row=row_idx, column=13, value=rapor.get("sehir", ""))
        ws.cell(row=row_idx, column=14, value=rapor.get("aciklama", ""))
        created_at = rapor.get("created_at", "")
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        ws.cell(row=row_idx, column=15, value=created_at.strftime("%Y-%m-%d %H:%M") if created_at else "")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=raporlar_{len(raporlar)}_adet.xlsx"}
    )

@router.get("/export-all")
async def export_excel(current_user: dict = Depends(get_current_user)):
    """Tüm raporları Excel'e aktar"""
    raporlar = await db.raporlar.find({}, {"_id": 0}).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Raporlar"
    
    headers = [
        "Rapor No", "Ekipman Adı", "Kategori", "Firma", "Lokasyon",
        "Marka/Model", "Seri No", "Alt Kategori", "Periyot",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama", "Oluşturma Tarihi"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, rapor in enumerate(raporlar, 2):
        ws.cell(row=row_idx, column=1, value=rapor.get("rapor_no", ""))
        ws.cell(row=row_idx, column=2, value=rapor.get("ekipman_adi", ""))
        ws.cell(row=row_idx, column=3, value=rapor.get("kategori", ""))
        ws.cell(row=row_idx, column=4, value=rapor.get("firma", ""))
        ws.cell(row=row_idx, column=5, value=rapor.get("lokasyon", ""))
        ws.cell(row=row_idx, column=6, value=rapor.get("marka_model", ""))
        ws.cell(row=row_idx, column=7, value=rapor.get("seri_no", ""))
        ws.cell(row=row_idx, column=8, value=rapor.get("alt_kategori", ""))
        ws.cell(row=row_idx, column=9, value=rapor.get("periyot", ""))
        ws.cell(row=row_idx, column=10, value=rapor.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=11, value=rapor.get("uygunluk", ""))
        ws.cell(row=row_idx, column=12, value=rapor.get("aciklama", ""))
        created_at = rapor.get("created_at", "")
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        ws.cell(row=row_idx, column=13, value=created_at.strftime("%Y-%m-%d %H:%M") if created_at else "")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=raporlar.xlsx"}
    )

@router.get("/template")
async def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor Şablonu"
    
    headers = [
        "Şehir", "Ekipman Adı", "Kategori", "Firma", "Lokasyon", "Marka/Model",
        "Seri No", "Alt Kategori", "Periyot", "Geçerlilik Tarihi",
        "Uygunluk", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    examples = [
        ["İstanbul", "Asansör A1", "Asansör", "ABC Firma", "İstanbul Ofis", "Otis 2000",
         "SN12345", "Yolcu Asansörü", "6 Aylık", "2025-12-31", "Uygun", "Örnek açıklama"],
        ["Ankara", "Hidrofor H2", "Basınçlı Kaplar", "XYZ Ltd", "Ankara Fabrika", "Grundfos",
         "SN67890", "Hidrofor", "3 Aylık", "2025-09-30", "Uygun Değil", ""],
        ["İzmir", "Forklift F3", "Forklift", "Test Firma", "İzmir Depo", "Toyota", "", "", "", "", "", ""]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    ws_cities = wb.create_sheet("Şehir Listesi")
    ws_cities.cell(row=1, column=1, value="Geçerli Şehirler (Büyük/küçük harf önemli değil)")
    ws_cities.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    for idx, sehir in enumerate(SEHIRLER, 2):
        ws_cities.cell(row=idx, column=1, value=sehir["isim"])
    
    ws_cities.column_dimensions['A'].width = 25
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rapor_sablonu.xlsx"}
    )

@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    proje_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Excel içe aktarma yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    proje_adi = proje.get("proje_adi", "")
    
    content = await file.read()
    excel_file = io.BytesIO(content)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            
            try:
                def get_cell(index):
                    return row[index] if index < len(row) and row[index] is not None else None
                
                sehir = str(get_cell(0)) if get_cell(0) else None
                ekipman_adi = str(get_cell(1)) if get_cell(1) else ""
                kategori = str(get_cell(2)) if get_cell(2) else ""
                firma = str(get_cell(3)) if get_cell(3) else ""
                lokasyon = str(get_cell(4)) if get_cell(4) else None
                marka_model = str(get_cell(5)) if get_cell(5) else None
                seri_no = str(get_cell(6)) if get_cell(6) else None
                alt_kategori = str(get_cell(7)) if get_cell(7) else None
                periyot = str(get_cell(8)) if get_cell(8) else None
                gecerlilik_tarihi = str(get_cell(9)) if get_cell(9) else None
                uygunluk = str(get_cell(10)) if get_cell(10) else None
                aciklama = str(get_cell(11)) if get_cell(11) else None
                
                if not ekipman_adi or not kategori or not firma:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik (Ekipman Adı, Kategori, Firma)")
                    continue
                
                if not sehir:
                    errors.append(f"Satır {row_idx}: Şehir alanı zorunludur")
                    continue
                
                def turkish_lower(text):
                    text = text.strip()
                    text = text.replace('İ', 'i').replace('I', 'ı')
                    return text.lower()
                
                def normalize_turkish(text):
                    turkish_map = str.maketrans('ıİiIğĞüÜşŞöÖçÇ', 'iiiigguussoocc')
                    return turkish_lower(text).translate(turkish_map)
                
                sehir_normalized = normalize_turkish(sehir)
                
                sehir_obj = None
                for s in SEHIRLER:
                    if normalize_turkish(s["isim"]) == sehir_normalized:
                        sehir_obj = s
                        sehir = s["isim"]
                        break
                
                if not sehir_obj:
                    errors.append(f"Satır {row_idx}: Geçersiz şehir - '{sehir}'")
                    continue
                
                rapor_no = await generate_rapor_no(sehir)
                
                rapor_data = {
                    "id": str(uuid.uuid4()),
                    "rapor_no": rapor_no,
                    "proje_id": proje_id,
                    "proje_adi": proje_adi,
                    "sehir": sehir,
                    "sehir_kodu": sehir_obj["kod"],
                    "ekipman_adi": ekipman_adi,
                    "kategori": kategori,
                    "alt_kategori": alt_kategori,
                    "firma": firma,
                    "lokasyon": lokasyon,
                    "marka_model": marka_model,
                    "seri_no": seri_no,
                    "periyot": periyot,
                    "gecerlilik_tarihi": gecerlilik_tarihi,
                    "uygunluk": uygunluk,
                    "aciklama": aciklama,
                    "durum": "Aktif",
                    "created_by": current_user["id"],
                    "created_by_username": current_user.get("username", current_user.get("email", "")),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.raporlar.insert_one(rapor_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
        
        return {
            "message": f"{imported_count} rapor başarıyla içe aktarıldı",
            "imported_count": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel işleme hatası: {str(e)}")
