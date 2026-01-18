from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
import io
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import IskeleBileseni, IskeleBileseniCreate
from routers.auth import get_current_user
from database import db

router = APIRouter(tags=["Iskele"])

# ==================== İSKELE BİLEŞEN ADLARI ====================

@router.get("/iskele-bilesen-adlari")
async def get_iskele_bilesen_adlari(current_user: dict = Depends(get_current_user)):
    bilesen_adlari = await db.iskele_bilesen_adlari.find({}, {"_id": 0}).to_list(1000)
    return bilesen_adlari

@router.post("/iskele-bilesen-adlari")
async def create_iskele_bilesen_adi(
    bilesen_adi: str,
    aciklama: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.iskele_bilesen_adlari.find_one({"bilesen_adi": bilesen_adi}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Bu bileşen adı zaten mevcut")
    
    bilesen_id = str(uuid.uuid4())
    bilesen_data = {
        "id": bilesen_id,
        "bilesen_adi": bilesen_adi,
        "aciklama": aciklama,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.iskele_bilesen_adlari.insert_one(bilesen_data)
    created = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    return created

@router.put("/iskele-bilesen-adlari/{bilesen_id}")
async def update_iskele_bilesen_adi(
    bilesen_id: str,
    bilesen_adi: str,
    aciklama: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Bileşen adı bulunamadı")
    
    update_data = {
        "bilesen_adi": bilesen_adi,
        "aciklama": aciklama
    }
    
    await db.iskele_bilesen_adlari.update_one({"id": bilesen_id}, {"$set": update_data})
    updated = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    return updated

@router.delete("/iskele-bilesen-adlari/{bilesen_id}")
async def delete_iskele_bilesen_adi(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.iskele_bilesen_adlari.delete_one({"id": bilesen_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bileşen adı bulunamadı")
    
    return {"message": "Bileşen adı silindi"}

# ==================== İSKELE BİLEŞENLERİ ====================

@router.get("/iskele-bilesenleri")
async def get_iskele_bilesenleri(
    current_user: dict = Depends(get_current_user),
    limit: int = 500
):
    query = {}
    if current_user.get("role") == "viewer" and current_user.get("firma_adi"):
        query["firma_adi"] = current_user.get("firma_adi")
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(limit)
    return bilesenleri

@router.post("/iskele-bilesenleri")
async def create_iskele_bileseni(
    bilesen: IskeleBileseniCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni ekleme yetkiniz yok")
    
    proje = await db.projeler.find_one({"id": bilesen.proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    bilesen_id = str(uuid.uuid4())
    bilesen_data = {
        **bilesen.model_dump(),
        "id": bilesen_id,
        "proje_adi": proje.get("proje_adi", ""),
        "iskele_periyodu": "6 Aylık",
        "created_by": current_user["id"],
        "created_by_username": current_user.get("username", current_user.get("email", "")),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.iskele_bilesenleri.insert_one(bilesen_data)
    
    created = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    return created

@router.get("/iskele-bilesenleri/{bilesen_id}")
async def get_iskele_bileseni(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    bilesen = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    if not bilesen:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    return bilesen

@router.put("/iskele-bilesenleri/{bilesen_id}")
async def update_iskele_bileseni(
    bilesen_id: str,
    bilesen_update: IskeleBileseniCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni güncelleme yetkiniz yok")
    
    existing = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    
    update_data = bilesen_update.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["iskele_periyodu"] = "6 Aylık"
    
    await db.iskele_bilesenleri.update_one(
        {"id": bilesen_id},
        {"$set": update_data}
    )
    
    updated = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    return updated

@router.delete("/iskele-bilesenleri/{bilesen_id}")
async def delete_iskele_bileseni(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni silme yetkiniz yok")
    
    result = await db.iskele_bilesenleri.delete_one({"id": bilesen_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    
    return {"message": "İskele bileşeni silindi"}

@router.post("/iskele-bilesenleri/bulk-delete")
async def bulk_delete_iskele_bilesenleri_route(
    bilesen_ids: List[str],
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni silme yetkiniz yok")
    
    if not bilesen_ids:
        raise HTTPException(status_code=400, detail="Silinecek bileşen ID'leri belirtilmedi")
    
    result = await db.iskele_bilesenleri.delete_many({"id": {"$in": bilesen_ids}})
    return {"message": f"{result.deleted_count} iskele bileşeni silindi", "deleted_count": result.deleted_count}

# ==================== İSKELE EXCEL ====================

@router.get("/iskele-bilesenleri/excel/export")
async def export_iskele_excel(current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user.get("role") == "viewer" and current_user.get("firma_adi"):
        query["firma_adi"] = current_user.get("firma_adi")
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "İskele Bileşenleri"
    
    headers = [
        "Bileşen Adı", "Malzeme Kodu", "Bileşen Adedi", "Firma Adı",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama", "Proje Adı"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, bilesen in enumerate(bilesenleri, 2):
        ws.cell(row=row_idx, column=1, value=bilesen.get("bileşen_adi", ""))
        ws.cell(row=row_idx, column=2, value=bilesen.get("malzeme_kodu", ""))
        ws.cell(row=row_idx, column=3, value=bilesen.get("bileşen_adedi", 0))
        ws.cell(row=row_idx, column=4, value=bilesen.get("firma_adi", ""))
        ws.cell(row=row_idx, column=5, value=bilesen.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=6, value=bilesen.get("uygunluk", ""))
        ws.cell(row=row_idx, column=7, value=bilesen.get("aciklama", ""))
        ws.cell(row=row_idx, column=8, value=bilesen.get("proje_adi", ""))
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=iskele_bilesenleri_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@router.get("/iskele-bilesenleri/excel/template")
async def download_iskele_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "İskele Bileşenleri Şablonu"
    
    headers = [
        "Bileşen Adı", "Malzeme Kodu", "Bileşen Adedi", "Firma Adı",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    examples = [
        ["Çelik Direk", "CD-001", 10, "ABC İnşaat", "2025-12-31", "Uygun", "Standart çelik direk"],
        ["Bağlantı Elemanı", "BE-002", 50, "XYZ Yapı", "2025-06-30", "Uygun", ""],
        ["Destek Parçası", "DP-003", 25, "Test Firma", "", "Uygun Değil", "Kontrol gerekli"]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=iskele_bilesenleri_sablonu.xlsx"}
    )

@router.post("/iskele-bilesenleri/excel/import")
async def import_iskele_excel(
    file: UploadFile = File(...),
    proje_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni içe aktarma yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    proje_adi = proje.get("proje_adi", "")
    
    content = await file.read()
    excel_file = io.BytesIO(content)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            
            try:
                def get_cell(index):
                    return row[index] if index < len(row) and row[index] is not None else None
                
                bilesen_adi = str(get_cell(0)) if get_cell(0) else ""
                malzeme_kodu = str(get_cell(1)) if get_cell(1) else ""
                bilesen_adedi_raw = get_cell(2)
                firma_adi = str(get_cell(3)) if get_cell(3) else ""
                gecerlilik_tarihi = str(get_cell(4)) if get_cell(4) else None
                uygunluk = str(get_cell(5)) if get_cell(5) else "Uygun"
                aciklama = str(get_cell(6)) if get_cell(6) else None
                
                if not bilesen_adi or not malzeme_kodu or not firma_adi:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik")
                    continue
                
                try:
                    bilesen_adedi = int(bilesen_adedi_raw) if bilesen_adedi_raw else 1
                    if bilesen_adedi < 1:
                        errors.append(f"Satır {row_idx}: Bileşen adedi en az 1 olmalıdır")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Satır {row_idx}: Bileşen adedi geçersiz")
                    continue
                
                bilesen_id = str(uuid.uuid4())
                bilesen_data = {
                    "id": bilesen_id,
                    "proje_id": proje_id,
                    "proje_adi": proje_adi,
                    "bileşen_adi": bilesen_adi,
                    "malzeme_kodu": malzeme_kodu,
                    "bileşen_adedi": bilesen_adedi,
                    "firma_adi": firma_adi,
                    "iskele_periyodu": "6 Aylık",
                    "gecerlilik_tarihi": gecerlilik_tarihi,
                    "uygunluk": uygunluk,
                    "aciklama": aciklama,
                    "gorseller": [],
                    "created_by": current_user["id"],
                    "created_by_username": current_user.get("username", current_user.get("email", "")),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.iskele_bilesenleri.insert_one(bilesen_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
                continue
        
        return {
            "message": f"{imported_count} iskele bileşeni başarıyla içe aktarıldı",
            "imported_count": imported_count,
            "errors": errors
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel dosyası işlenemedi: {str(e)}")
